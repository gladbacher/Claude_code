"""
Excel and JSON export for match and group results.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed — Excel export disabled. pip install openpyxl")


# ── Colour fills ─────────────────────────────────────────────────────────────
_GREEN = PatternFill("solid", fgColor="C6EFCE") if _EXCEL_AVAILABLE else None
_AMBER = PatternFill("solid", fgColor="FFEB9C") if _EXCEL_AVAILABLE else None
_RED   = PatternFill("solid", fgColor="FFC7CE") if _EXCEL_AVAILABLE else None
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79") if _EXCEL_AVAILABLE else None
_HEADER_FONT = Font(color="FFFFFF", bold=True) if _EXCEL_AVAILABLE else None


def _edge_fill(edge: float):
    if edge > 0.05:
        return _GREEN
    if edge > 0.02:
        return _AMBER
    return _RED


def export_json(data: Any, path: Path) -> None:
    """Write arbitrary data as JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, default=str)
    logger.info("JSON exported to %s", path)


def export_group_excel(
    group: str,
    match_results: dict[tuple[str, str], Any],
    group_sim: Any,
    output_path: Path | None = None,
) -> Path | None:
    """
    Write a 3-sheet Excel workbook for a group:
      Sheet 1 — Fixtures & main markets
      Sheet 2 — Qualification probabilities
      Sheet 3 — AH probability grid
    """
    if not _EXCEL_AVAILABLE:
        logger.error("openpyxl required for Excel export")
        return None

    if output_path is None:
        output_path = Path(f"wc2026_group_{group}.xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Fixtures ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"Group {group} Fixtures"

    headers = ["Home", "Away", "λ Home", "λ Away",
               "Home Win%", "Draw%", "Away Win%",
               "Over 2.5%", "Under 2.5%", "BTTS Yes%", "BTTS No%"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, ((home, away), res) in enumerate(match_results.items(), 2):
        values = [
            home, away,
            round(res["lambda_home"], 2), round(res["lambda_away"], 2),
            f"{res['home_win']*100:.1f}%",
            f"{res['draw']*100:.1f}%",
            f"{res['away_win']*100:.1f}%",
            f"{res['over_25']*100:.1f}%",
            f"{(1-res['over_25'])*100:.1f}%",
            f"{res['btts']*100:.1f}%",
            f"{(1-res['btts'])*100:.1f}%",
        ]
        for col, v in enumerate(values, 1):
            ws1.cell(row=row_idx, column=col, value=v)

    # ── Sheet 2: Qualification ───────────────────────────────────────────────
    ws2 = wb.create_sheet(f"Group {group} Qualification")
    q_headers = ["Team", "Qualify%", "Win Group%", "Third Place%", "Avg Points"]
    for col, h in enumerate(q_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    teams_sorted = sorted(
        group_sim["qualification"].keys(),
        key=lambda t: group_sim["qualification"][t],
        reverse=True,
    )
    for row_idx, team in enumerate(teams_sorted, 2):
        ws2.cell(row=row_idx, column=1, value=team)
        ws2.cell(row=row_idx, column=2, value=f"{group_sim['qualification'][team]*100:.1f}%")
        ws2.cell(row=row_idx, column=3, value=f"{group_sim['win_group'][team]*100:.1f}%")
        ws2.cell(row=row_idx, column=4, value=f"{group_sim['third_place'][team]*100:.1f}%")
        ws2.cell(row=row_idx, column=5, value=round(group_sim["points_avg"][team], 2))

    # ── Sheet 3: AH grid ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet(f"Group {group} AH")
    ah_headers = ["Home", "Away", "Line", "Home Cover%", "Push%", "Away Cover%"]
    for col, h in enumerate(ah_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_idx = 2
    from .model import AH_LINES
    for (home, away), res in match_results.items():
        for line in AH_LINES:
            ws3.cell(row=row_idx, column=1, value=home)
            ws3.cell(row=row_idx, column=2, value=away)
            ws3.cell(row=row_idx, column=3, value=line)
            ws3.cell(row=row_idx, column=4, value=f"{res['ah_home'][line]*100:.1f}%")
            ws3.cell(row=row_idx, column=5, value=f"{res['ah_push'][line]*100:.1f}%")
            ws3.cell(row=row_idx, column=6, value=f"{res['ah_away'][line]*100:.1f}%")
            row_idx += 1

    wb.save(output_path)
    logger.info("Excel exported to %s", output_path)
    return output_path
